from openpyxl import load_workbook


def getRowCount(file, sheetName):
    workbook = load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColomnCount(file, sheetName):
    workbook = load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_colomn


def readDate(file, sheetName, row_num, column_num):
    workbook = load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=row_num, column=column_num).value


def writeDate(file, sheetName, row_num, column_num, data):
    workbook = load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=row_num, column=column_num).value = data
    workbook.save(file)
